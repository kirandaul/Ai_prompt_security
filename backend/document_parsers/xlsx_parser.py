"""
XLSX (Excel) document parser using openpyxl library.
Extracts text from cells with sheet/row/column tracking.
"""

from typing import Dict, List, Any
from .base_parser import BaseDocumentParser

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None


class XLSXParser(BaseDocumentParser):
    """Parse XLSX (Excel) documents."""
    
    async def parse(self, file_content: bytes, filename: str) -> Dict[str, Any]:
        """
        Parse XLSX and extract cell values with location.
        
        Returns:
            {
                "success": bool,
                "filename": str,
                "file_type": "xlsx",
                "pages": [
                    {
                        "sheet": "Sheet1",
                        "cell": "A1",
                        "text": "...",
                        "length": int
                    },
                    ...
                ],
                "total_pages": int,  # Number of cells with content
                "text": "combined text",
                "metadata": {...}
            }
        """
        
        if not load_workbook:
            return {
                "success": False,
                "error": "openpyxl library not installed",
                "filename": filename,
                "file_type": "xlsx"
            }
        
        try:
            import io
            workbook = load_workbook(io.BytesIO(file_content), data_only=True)
            
            pages = []
            all_text = []
            
            # Extract from all sheets
            for sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
                
                for row in worksheet.iter_rows(values_only=False):
                    for cell in row:
                        if cell.value and str(cell.value).strip():
                            cell_text = str(cell.value).strip()
                            pages.append({
                                "sheet": sheet_name,
                                "cell": cell.coordinate,
                                "text": cell_text,
                                "length": len(cell_text)
                            })
                            all_text.append(f"--- {sheet_name}[{cell.coordinate}] ---\n{cell_text}")
            
            # Extract metadata
            metadata = {}
            try:
                if workbook.properties:
                    props = workbook.properties
                    metadata = {
                        "title": props.title or "",
                        "subject": props.subject or "",
                        "author": props.author or "",
                        "comments": props.comments or "",
                        "category": props.category or "",
                        "created": str(props.created) if props.created else "",
                        "modified": str(props.modified) if props.modified else "",
                    }
            except:
                pass
            
            metadata["sheet_names"] = workbook.sheetnames
            
            return {
                "success": True,
                "filename": filename,
                "file_type": "xlsx",
                "pages": pages,
                "total_pages": len(pages),
                "text": "\n".join(all_text),
                "metadata": metadata
            }
            
        except Exception as e:
            return {
                "success": False,
                "filename": filename,
                "file_type": "xlsx",
                "error": str(e),
                "pages": [],
                "total_pages": 0,
                "text": ""
            }
